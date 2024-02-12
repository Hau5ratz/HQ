import xlwings as xw
import os
import sys
from datetime import datetime
import datetime as dt
import pickle
import time as t
from pprint import pp
import openpyxl
import json
import sqlite3
import logging
import csv
import time

BCODE = '''Sub get_cust_md(EID As String, Password As String, mdx As String)

Dim sts As String

Sheet1.Visible = True
Sheet1.Activate

sts = HypUIConnect("Q", EID, Password, "")
sts = HypExecuteQuery(Empty, mdx)

End Sub



query's are faster in dollars and then should be converted for euros
'''
QO = "([%s],[%s],[%s],[%s])"
with open("iter.obj","rb") as file:
    D = pickle.load(file)
GEOS = sorted(list(set([x[0] for x in D[0]])))
ACCOUNTS = sorted(D[1])
YNS = ['y','n','Y','N','yes','no','Yes','No','YES','NO']
fn = "COS template.xlsm"      
mn = "get_cust_md"
cy = datetime.now().strftime("%Y")
py = str(int(cy)-1)
months = [datetime(int(cy),x,1).strftime("%b") for x in range(1,13)]
# Template goes {([year],[month]),...}
template = "select {%s} on columns, non Empty {%s} on rows from [ccosrpt].[report] where ( [actual_final],  [tot_product], [source], [cmci], [usd - reporting] )" #formerly usd was reporting_eur
password = "brave salmon butcher"
c = 0
logging.basicConfig(filename='DB.log',filemode='a', encoding='utf-8', level=logging.DEBUG)

SQL_DB_path = "C:\\Users\\e065057\\OneDrive - Mastercard\\FINANCE\\Database\\Main.db"

def sql2csv():
    conn = create_connection()
    cur = conn.cursor()
    data = cur.execute("SELECT * FROM FF")
    with open('db.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['geo', 'g_code', 'div', 'acc', 'month', 'year', 'value', 'alias', 'iss_acq', 'map_one', 'map_two', 'map_three', 'exclude', 'percent'])
        writer.writerows(data)
    print('Finished writing to CSV')
    
def argue(query):
    conn = create_connection()
    cur = conn.cursor()
    data = cur.execute(query)
    return cur.fetchall()
    
def table_columns(db, table_name):
    if db == None:
        conn = create_connection()
        curs = conn.cursor()
    else:    
        curs = db.cursor()
    sql = "select * from %s where 1=0;" % table_name
    curs.execute(sql)
    return [d[0] for d in curs.description]

def create_row(conn,table, insert):
    """
    Create a new project into the projects table
    :param conn:
    :param project:
    :return: project id
    """
    sql = f''' INSERT INTO {table}(geo,g_code,div,acc,month,year,value,alias)
              VALUES(?,?,?,?,?,?,?,?) '''
    cur = conn.cursor()
    cur.execute(sql, insert)
    #conn.commit()
    return cur.lastrowid

def create_connection(db_file=SQL_DB_path):
    """ create a database connection to the SQLite database
        specified by the db_file
    :param db_file: database file
    :return: Connection object or None
    """
    conn = None
    try:
        conn = sqlite3.connect(db_file)
    except Error as e:
        print(e)
        logging.error(f"Error {e} occured") 

    return conn

def sql_up():
    '''
    Uploads data to sql db
    
    ff schema:
    [(0, 'geo', 'VARCHAR(255)', 1, None, 0),
    (1, 'g_code', 'VARCHAR(255)', 0, None, 1),
    (2, 'div', 'VARCHAR(255)', 1, None, 0),
    (3, 'acc', 'VARCHAR(255)', 1, None, 0),
    (4, 'month', 'INTEGER', 1, None, 0),
    (5, 'year', 'INTEGER', 1, None, 0),
    (6, 'alias', 'VARCHAR(255)', 1, None, 0),
    (7, 'iss_acq', 'INTEGER', 1, '0', 0),
    (8, 'map_one', 'VARCHAR(255)', 0, None, 0),
    (9, 'map_two', 'VARCHAR(255)', 0, None, 0),
    (10, 'map_three', 'VARCHAR(255)', 0, None, 0),
    (11, 'exclude', 'INTEGER', 1, '0', 0),
    (12, 'percent', 'INTEGER', 1, '100', 0)]
    '''
    conn = create_connection(SQL_DB_path)
    
    with open('HYPD.obj','rb') as file:
        d = pickle.load(file)
    
    for r in d['Data']['2023-10-19']:
        #needs values: (geo,g_code,div,acc,month,year,value,alias) 

        x = 0
        for c in r[4:]:
            p = r[:4]

            p += [months[x%12]]
            if x < 12:
                p += ['2022']
            else:
                p += ['2023']

            p += [c]
            p += [r[1][15:]]
            x += 1
            #pp(p)
            try:
                create_row(conn, "FF", tuple(p))
                print(f"Record: {d['Data']['2023-10-19'].index(r)} out of {len(d['Data']['2023-10-19'])} submitted, {int((d['Data']['2023-10-19'].index(r)/len(d['Data']['2023-10-19']))*100)}% completed", end="\r")
            except:
                logging.info(f'Offending record is {p}')
        

    #Need to log submission errors
    print("Commencing commit")
    conn.commit()
        
    #cursor.execute("PRAGMA table_info(FF);")
    #pp(cursor.fetchall())
    
    

def xlin():
    '''
    This function formats the HYPD pickle data into the Matteo's DB.xlsx
    '''
    rr = 2
    with open('HYPD.obj','rb') as file:
        d = pickle.load(file)
        
    wb =  openpyxl.load_workbook('DB.xlsx')
    ws = wb['DB-Actuals']
    for r in d['Data']['2023-10-19']:
        p = r[:4]
        x = 0
        for c in r[4:]:
            if x < 12:
                i = ['2022']
            else:
                i = ['2023']
            i += [months[x%12]]
            i += p 
            for item in i:
                ws[f"{chr(65 + i.index(item))}{rr+d['Data']['2023-10-19'].index(r)}"].value = item
            ws[f"H{rr+d['Data']['2023-10-19'].index(r)}"].value = c
            rr += 1 
            x += 1

def y_n(msg):
    '''
    A simple Yes no logic tree for prompting the user questions
    '''
    while True:
        y_n = input(msg)
        if not y_n in YNS:
            print('Invalid response')
        else:
            if 'n' in y_n.lower():
                return False
            else:
                return True

def cls():
    '''
    Calls windows clear screen function
    '''
    #os.system('cls')
    pass

def get_lr(ws):
    '''
    Get's the last row of the worksheet
    '''
    r = 4
    while True:
        x = ws.range(r,1).value
        if x:
            r += 1
        else:
            lr = r
            break
    return lr
    
    
def lister(l):
    '''
    Reads out items of a given list with their respective indices
    then prompts the user to pick one which it returns
    '''
    innie = False
    while not innie:
        print('Please select an item: ')
        [print(f"{l.index(x)}- {x}") for x in l]
        innie = input("Item index?: ")
        try:
            outtie = l[int(innie)]
        except:
            print('\n This is response is not a listed item please try again')
            innie = False
    return outtie
    
def t_builder():
    t = ""
    y = []
    ms = []
    while True:
        print('Example range: \n2000-2021\n2023')
        years = input("Please input a year range: ")
        try:
            yl = [int(y) for y in years.split("-")]
            if len(yl) == 2:
                yl = [x for x in range(yl[0],yl[1]+1)]
            break
        except:
            print("This is not a valid year range please try again")
            
    ml = False
    if y_n("Would you like monthly data? (y/n): "):
        while True:
            print('Example range: \n1-12 (jan-dec)\n5-8 (may-aug)\n6 (june)')
            months = input("Please input a month range: ")
            try:
                ml = [int(m) for m in months.split("-")]
                if len(ml) == 2:
                    ml = [x for x in range(ml[0],ml[1]+1)]
                break
            except:
                print("This is not a valid month range please try again")
    for y in yl:
        if ml:
            for m in ml:
                t += f"([{y}],[{months[m-1]}]),"
        else:
            t += f"([{y}]),"
    return t[:-1]
    
def formatter(t,a,b):
    q = []
    for o in b:
        q += [QO%(a[0],a[1],o.split(";")[0],o.split(";")[1])]
    mdx = template%(t,",".join(q))
    mdx = mdx.replace("['","")
    mdx = mdx.replace("']","")
    mdx = mdx.replace("', '",",")
    return mdx
    
    
def Query_wrapper(eid, p, q=None):
    '''
    Simple text based user interface for the query function
    Takes an Employee ID
    password
    query 
    '''
    cus = []
    accs = []
    gc = True
    while gc:
        g = lister(GEOS)
        cls()
        CC = input("Please input a valid customer code: ")
        if 13 >= len(CC) >= 10:
            cus += [[g,CC]]
            if y_n("Would you like to add an additional customer? (y/n): "):
                continue
            break
        else:
            print('This is not a valid CC please try again')
    while True:
        cls()
        print('For instance: ')
        acc = lister(ACCOUNTS)
        accs += [acc]
        if y_n('Do you wish to add an additional subject of query (y/n)?: '):
            continue
        else:
            break
    cls()
    t = t_builder()
    print("Current query: ")
    for c in cus:    
        mdx = formatter(t, c, accs)
        logging.info("Submitting query..")
        logging.info("Query:\n")
        print(mdx)
        logging.info(mdx)
        Query("brave salmon butcher", mdx)
        logging.info("Query completed")
    

def Query(p, x, eid="e065057", w = True):
    data = []
    with xw.App(visible=False) as app:
        wb = xw.Book(fn)
        m = wb.macro(mn)
        ws = wb.sheets[0]
        ws.activate()
        try:
            m(eid, p, x)
        except Exception as ex:
            e_service(ex)
            #wb.close()
            logging(f"Offending query is:\n{x}")
            print(f"Offending query is:\n{x}")
            w = False
            #exit()
        lr = get_lr(ws)
        for r in range(3,lr):
            data += [[ws.range(r,c).value for c in range(1,29)]]
        wb.close()
        
    if not data[0][0]:
        data = data[1:]
    if data == []:
        w = False
        
    if w:
        # read db and write to db new record
        with open("HYPD.obj","rb") as file:
            d = pickle.load(file)
        logging.info(f"Data submitted:{data}")
        time = str(dt.date.today())
        if time in d['Data'].keys():
            d['Data'][time] += data
        else:
            d['Data'][time] = data
        #d['Args'][time] = args
        with open("HYPD.obj","wb") as file:
            pickle.dump(d,file)
        logging.info("Record written")
        print("Record written")
        return d
    else:
        logging.info('No data submitted')
        print('No data submitted')
        #exit()
        return data
    #wb.close()
        
def e_service(ex):
    '''
    Exception handeling function
    '''
    print("Exception %s has occured"%ex)
    #print("Args: %s"%ex.args)
    print("Please consult the query template provided below when submitting a query")
    print(template%("Period of inquery as:\n\n([year],[month]),...}", "Geography code", "Customer code","Division","Revenue Account"))
    
        
    
def pop():
    '''
    Pop iterates through all possible combinations of division account parings 
    and then submits the combinations as a per customer query to hyperion
    '''
    q = []
    c = 0
    cs = list(set(["~".join(x) for x in D[0]]))
    
    for customer in [x.split('~') for x in cs]:
        c += 1
        t1 = time.time()
        print(f"Current query customer: {customer[1][:13]} progress is {int((c/len(cs))*100)}%", end="\r")
        # d is structured as [[country,customer],[division:accs]]
        with open("HYPD.obj","rb") as file:
            d2 = pickle.load(file)
            
        if str(dt.date.today()) in list(d2['Data'].keys()):
            if customer[1] in [x[1] for x in d2['Data'][str(dt.date.today())]]:
                print(f'Skipping customer {customer[1]}')
                continue
        t = ""
        c = False
        for y in [py,cy]:
            for m in months:
                if c:
                    t += ","
                t += f"([{y}],[{m}])"
                c = True
        mdx = formatter(t, customer, D[1])
        # test the query

        # Submit the query 
        Query("whatajoke1%", mdx)
        t2 = time.time()
        logging.info(f'Query resolved in {int(t2-t1)} seconds')
        #print("Query resolved")
    
            
            
def ff_up():
    '''
    Uploads stored Hyperion data into the FF
    '''
    wb =  openpyxl.load_workbook('MCDS Dashboard - August 2023 Financials - Edit.xlsm')
    ws = wb['Data']
    with open('HYPD.obj','rb') as file:
        d = pickle.load(file)
    d = d['Data'][str(dt.date.today())]
    letters = [chr(64+x) for x in range(1,27)] + ['AA','AB']
    for r in d:
        for l in r:
            ws[f"{letters[r.index(l)]}{2+d.index(r)}"].value = l
    wb.save('MCDS Dashboard - August 2023 Financials - Edit.xlsm')
    print('Workbook saved')
        
    
    
    
    
            
if __name__ == "__main__":
    cls()
    if sys.argv[-1] == "G":
        pop()
    elif "Q" in sys.argv:
        Query('brave salmon butcher',input("Query: "), "e065057")
    else:
        if len(sys.argv) <= 2:
            print("Needs EID and password as consequtive arguements")
            exit()
        else:
            Query_wrapper(sys.argv[-3], sys.argv[-2], sys.argv[-1] if len(sys.argv) >= 3 else False)
            
            
'''
from argparse import ArgumentParser
import matplotlib.pyplot as plt
import pandas as pd


def get_filepaths():
    """Collect excel file paths from the terminal."""

    parser = ArgumentParser(description="Plot excel files")
    parser.add_argument(
        "filepath",
        type=str,
        nargs="+",
        help="provide the full path of your excel files",
    )
    filepaths = parser.parse_args()
    return filepaths


def plot_excels(filepaths):
    """Load and plot one aspect of the excel files."""

    # load the excel files that we got from the provided arguments
    for path in filepaths.filepath:
        df = pd.read_excel(path)
        df.plot(kind="line", title="Per unit price & number of units")
        plt.show()


if __name__ == "__main__":
    filepaths = get_filepaths()
    plot_excels(filepaths)
'''
